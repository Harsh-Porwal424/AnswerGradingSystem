import os
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import openpyxl
import openai
import requests
from fuzzywuzzy import fuzz
import google.generativeai as genai
import threading
import time
import signal

# ---------------- Global Variables ----------------
ans = 0
text = ''
kt = 0
cm = 0
gm = 0
fr = 0
g = 0
file = ""
fileQ = ""
Qtext = ""
frf = 1
ktf = 0
cmf = 0
gmf = 0

strans = ""
ansl = []
keyword = []

genai.configure(api_key='AIzaSyDM5zj3dyNoMILBEYVNy00w8WKRdsLHy9U')

# Common words to ignore when extracting keywords
com = [
    "and", "that", "the", "for", "it", "it's", "was", "his", "who", "work",
    "used", "way", "also", "by", "can", "which", "as", "known", "then", "if",
    "between", "through", "another", "", "or", "my", "in", "from", "a", "any",
    "on", "combination", "to", "into", "is", "of", "it", "a", "each", "both"
]



# ---------------- OCR Functions ----------------
def extract_text_from_image(image_path: str) -> str:
    """Extracts text from handwritten image using Gemini API."""
    model = genai.GenerativeModel('gemini-1.5-flash')
    try:
        response = model.generate_content([
            "Extract all text from this handwritten answer sheet. "
            "Return only the raw text without any formatting. "
            "Preserve original line breaks and punctuation.",
            genai.upload_file(image_path)
        ])
        return response.text
    except Exception as e:
        raise RuntimeError(f"OCR Error: {str(e)}")

# ---------------- Utility Functions ----------------
def load_words():
    """Load valid English words from file."""
    try:
        with open('/Users/harshporwal/Desktop/MAIN/AI_NLP/words_alpha.txt', 'r') as word_file:
            valid_words = set(word_file.read().split())
        return valid_words
    except Exception as e:
        print("Error loading word list:", e)
        return set()

def openmyfile(file_name):
    """
    Loads the questions file for the given subject and extracts:
      - Questions text (Qtext_dict)
      - Weighting factors for each question (factors_dict)
      - Lists of sample answers (ansl_dict)
    """
    global fileQ, Qtext_dict, factors_dict, ansl_dict
    fileQ = file_name
    # Reset global dictionaries
    Qtext_dict = {}
    factors_dict = {}
    ansl_dict = {}
    
    print(f"Loading questions from file: {file_name}")  # Debug print
    
    # Adjust the path as needed for your system
    loc = f"/Users/harshporwal/Desktop/MAIN/AI_NLP/Questions/{fileQ}.xlsx"
    print(f"Full path: {loc}")  # Debug print
    
    if os.path.exists(loc):
        print(f"File exists, loading workbook...")  # Debug print
        wb = openpyxl.load_workbook(loc)
        print(f"Available sheets: {wb.sheetnames}")  # Debug print
        
        # Process each sheet (Q1 to Q5)
        for sheet_name in ['Q1', 'Q2', 'Q3', 'Q4', 'Q5']:
            if sheet_name in wb.sheetnames:
                print(f"Processing sheet: {sheet_name}")  # Debug print
                sheet = wb[sheet_name]
                # Store question text
                Qtext_dict[sheet_name] = sheet.cell(row=4, column=1).value
                # Store factors
                factors_dict[sheet_name] = {
                    'frf': sheet.cell(row=2, column=2).value,
                    'ktf': sheet.cell(row=2, column=3).value,
                    'cmf': sheet.cell(row=2, column=4).value,
                    'gmf': sheet.cell(row=2, column=5).value
                }
                # Process answer key for this question
                ansl_dict[sheet_name] = ans_key(sheet)
                print(f"Loaded question text: {Qtext_dict[sheet_name][:50]}...")  # Debug print
            else:
                print(f"Sheet {sheet_name} not found")  # Debug print
    else:
        print(f"Error: File '{loc}' not found.")
    
    print(f"Loaded {len(Qtext_dict)} questions")  # Debug print
    print(f"Question texts: {list(Qtext_dict.keys())}")  # Debug print

def ans_key(sheet):
    """
    Processes the answer key sheet and returns a dictionary containing:
    - sample_answers: list of sample answers
    - keywords: list of extracted keywords
    """
    answers = []
    keywords = set()

    # Collect sample answers from the sheet (from row 6 onward)
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        t = row[0]
        if t is not None:
            answers.append(t)

    # Extract keywords from the sample answers
    for a in answers:
        words_in_a = a.split()
        for word in words_in_a:
            word = word.lower()
            # Avoid self-comparison and common words
            for check in answers:
                if a == check:
                    continue
                else:
                    for x in check.split():
                        x = x.lower()
                        if x == word and x not in com:
                            keywords.add(x)

    return {
        'sample_answers': answers,
        'keywords': list(keywords)
    }

def get_available_subjects():
    """Get list of available subjects from the Questions folder."""
    questions_dir = "/Users/harshporwal/Desktop/MAIN/AI_NLP/Capstone Final/Questions"
    subjects = []
    
    # List all Excel files in the Questions directory
    for file in os.listdir(questions_dir):
        if file.endswith('.xlsx') and file != 'Scoring_Factors.xlsx':  # Exclude non-subject files
            # Remove the .xlsx extension to get the subject name
            subject = os.path.splitext(file)[0]
            subjects.append(subject)
    
    return sorted(subjects)  # Return sorted list of subjects

# ---------------- UI Classes ----------------
class LoginWindow(ttk.Frame):
    """The first window for login and subject selection."""
    def __init__(self, master):
        super().__init__(master, padding="20 20 20 20")
        self.master = master
        self.master.title("Automatic Answer Checker")
        self.grid(sticky="NSEW")
        self.create_widgets()
        # Initialize global variables
        global Qtext_dict, factors_dict, ansl_dict
        Qtext_dict = {}
        factors_dict = {}
        ansl_dict = {}
        
        # Get available subjects
        available_subjects = get_available_subjects()
        if available_subjects:  # If there are any subjects available
            # Set default subject to first available subject
            self.subject_var.set(available_subjects[0])
            openmyfile(available_subjects[0])
        else:
            messagebox.showerror("Error", "No subject files found in Questions directory!")

    def create_widgets(self):
        # Username label and entry
        ttk.Label(self, text="Username:", font=("Helvetica", 11)).grid(row=0, column=0, sticky="W", pady=5)
        self.username_entry = ttk.Entry(self, width=30)
        self.username_entry.grid(row=0, column=1, pady=5)
        self.username_entry.focus()

        # Password label and entry
        ttk.Label(self, text="Password:", font=("Helvetica", 11)).grid(row=1, column=0, sticky="W", pady=5)
        self.password_entry = ttk.Entry(self, width=30, show="*")
        self.password_entry.grid(row=1, column=1, pady=5)

        # Subject selection using Combobox
        ttk.Label(self, text="Subject:", font=("Helvetica", 11)).grid(row=2, column=0, sticky="W", pady=5)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(self, textvariable=self.subject_var, state="readonly", values=get_available_subjects())
        self.subject_combo.grid(row=2, column=1, pady=5)
        self.subject_combo.bind('<<ComboboxSelected>>', self.subject_changed)

        # Submit button
        self.submit_button = ttk.Button(self, text="Submit", command=self.check_credentials)
        self.submit_button.grid(row=3, column=1, pady=15, sticky="E")

    def subject_changed(self, event):
        selected = self.subject_var.get()
        print(f"\nSubject changed to: {selected}")  # Debug print
        openmyfile(selected)
        print(f"After loading, available questions: {list(Qtext_dict.keys())}")  # Debug print

    def check_credentials(self):
        user = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        # For demo purposes the credentials are hardcoded as 'admin'
        if user == "admin" and password == "admin":
            print("\nLogin successful, opening question window...")  # Debug print
            print(f"Current questions loaded: {list(Qtext_dict.keys())}")  # Debug print
            # Open the question window and hide the login window
            QuestionWindow(self.master)
            self.master.withdraw()
        else:
            messagebox.showerror("Login Error", "Please enter the correct credentials.")


# ---------------- Modified UI Classes ----------------
class QuestionWindow(tk.Toplevel):
    """Window for displaying multiple questions and answer inputs with OCR upload capability."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Question Paper")
        self.geometry("1000x800+200+50")
        
        print("Initializing QuestionWindow...")  # Debug print
        
        # Create main scrollable canvas
        self.canvas = tk.Canvas(self)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Pack scrollbar and canvas
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        self.answer_widgets = {}  # Store widgets for each question
        self.answers = {}  # Store answers for each question
        self.processing = {}  # Track OCR processing status
        
        print(f"Available questions in Qtext_dict: {list(Qtext_dict.keys())}")  # Debug print
        
        self.create_widgets()

    def create_widgets(self):
        print("Creating question widgets...")  # Debug print
        
        # Create frames for each question
        for q_num in range(1, 6):
            sheet_name = f'Q{q_num}'
            print(f"Processing question {sheet_name}")  # Debug print
            if sheet_name in Qtext_dict:
                print(f"Found question {sheet_name} in Qtext_dict")  # Debug print
                self.create_question_frame(sheet_name, q_num)
            else:
                print(f"Question {sheet_name} not found in Qtext_dict")  # Debug print
        
        # Add submit all button at the bottom
        submit_frame = ttk.Frame(self.scrollable_frame)
        submit_frame.pack(fill="x", padx=10, pady=20)
        
        submit_all_btn = ttk.Button(
            submit_frame, 
            text="Submit All Answers", 
            command=self.submit_all_answers
        )
        submit_all_btn.pack(side="right", padx=10)
        print("Finished creating widgets")  # Debug print

    def create_question_frame(self, sheet_name, q_num):
        print(f"Creating frame for question {sheet_name}")  # Debug print
        # Create frame for this question
        q_frame = ttk.LabelFrame(
            self.scrollable_frame, 
            text=f"Question {q_num}", 
            padding="10 10 10 10"
        )
        q_frame.pack(fill="x", padx=10, pady=5, expand=True)
        
        # Question text
        question_text = Qtext_dict[sheet_name]
        print(f"Question text for {sheet_name}: {question_text[:50]}...")  # Debug print
        
        ttk.Label(
            q_frame, 
            text=question_text,
            wraplength=900,
            font=("Helvetica", 11)
        ).pack(fill="x", pady=5)
        
        # Answer frame
        answer_frame = ttk.Frame(q_frame)
        answer_frame.pack(fill="x", pady=5)
        
        # Text widget for answer
        answer_text = tk.Text(
            answer_frame, 
            height=8, 
            width=80, 
            wrap="word",
            font=("Helvetica", 11)
        )
        answer_text.pack(side="left", padx=5)
        
        # Button frame
        btn_frame = ttk.Frame(q_frame)
        btn_frame.pack(fill="x", pady=5)
        
        # Upload button
        upload_btn = ttk.Button(
            btn_frame,
            text="Upload Handwritten Answer",
            command=lambda sn=sheet_name: self.upload_image(sn)
        )
        upload_btn.pack(side="left", padx=5)
        
        # Store widgets for this question
        self.answer_widgets[sheet_name] = {
            'text': answer_text,
            'upload_btn': upload_btn
        }
        self.processing[sheet_name] = False

    def upload_image(self, sheet_name):
        """Handle image upload and text extraction for a specific question."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.bmp *.tiff")]
        )
        if file_path:
            self.start_ocr_processing(sheet_name, file_path)

    def start_ocr_processing(self, sheet_name, image_path):
        """Start OCR processing in a separate thread for a specific question."""
        self.processing[sheet_name] = True
        widgets = self.answer_widgets[sheet_name]
        widgets['upload_btn'].config(state=tk.DISABLED, text="Processing...")
        
        def ocr_thread():
            try:
                extracted_text = extract_text_from_image(image_path)
                self.after(0, self.update_answer_text, sheet_name, extracted_text)
            except Exception as e:
                self.after(0, messagebox.showerror, "OCR Error", str(e))
            finally:
                self.after(0, self.complete_ocr_processing, sheet_name)
                
        threading.Thread(target=ocr_thread, daemon=True).start()

    def update_answer_text(self, sheet_name, text):
        """Update answer text widget with OCR results for a specific question."""
        widgets = self.answer_widgets[sheet_name]
        widgets['text'].delete("1.0", tk.END)
        widgets['text'].insert("1.0", text)

    def complete_ocr_processing(self, sheet_name):
        """Cleanup after OCR processing for a specific question."""
        self.processing[sheet_name] = False
        widgets = self.answer_widgets[sheet_name]
        widgets['upload_btn'].config(
            state=tk.NORMAL, 
            text="Upload Handwritten Answer"
        )

    def submit_all_answers(self):
        """Handle submission of all answers."""
        # Collect all answers
        all_answers = {}
        for sheet_name, widgets in self.answer_widgets.items():
            text = widgets['text'].get("1.0", tk.END).strip()
            if not text:
                messagebox.showerror(
                    "Input Error", 
                    f"Please enter or upload an answer for Question {sheet_name[1]}"
                )
                return
            elif len(text.split()) < 5:
                messagebox.showerror(
                    "Input Error", 
                    f"Answer for Question {sheet_name[1]} is too short! Minimum 5 words required."
                )
                return
            all_answers[sheet_name] = text
        
        # Store answers and proceed to report
        self.answers = all_answers
        ReportWindow(self)
        self.withdraw()


class ReportWindow(tk.Toplevel):
    """Window displaying the evaluation report for all questions."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Evaluation Report")
        self.geometry("800x600+400+100")
        self.configure(bg="#f7f7f7")
        
        # Get answers from parent window
        self.answers = parent.answers
        
        # Store evaluation results
        self.results = {}
        
        # Create scrollable frame
        self.create_scrollable_frame()
        self.create_widgets()
        
        # Evaluate all answers
        self.evaluate_all_answers()

    def create_scrollable_frame(self):
        # Create canvas with scrollbar
        self.canvas = tk.Canvas(self)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Pack scrollbar and canvas
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

    def create_widgets(self):
        # Title
        ttk.Label(
            self.scrollable_frame,
            text="Evaluation Results",
            font=("Helvetica", 16, "bold")
        ).pack(pady=10)
        
        # Frame for total marks
        self.total_frame = ttk.LabelFrame(
            self.scrollable_frame,
            text="Total Score",
            padding="10 10 10 10"
        )
        self.total_frame.pack(fill="x", padx=10, pady=5)
        
        # Individual question results will be added during evaluation
        self.question_frames = {}
        for sheet_name in sorted(self.answers.keys()):
            frame = ttk.LabelFrame(
                self.scrollable_frame,
                text=f"Question {sheet_name[1]}",
                padding="10 10 10 10"
            )
            frame.pack(fill="x", padx=10, pady=5)
            self.question_frames[sheet_name] = frame
        
        # Buttons
        button_frame = ttk.Frame(self.scrollable_frame)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Button(
            button_frame,
            text="View Detailed Report",
            command=self.open_detailed_report
        ).pack(side="left", padx=5)
        
        ttk.Button(
            button_frame,
            text="Close",
            command=self.close_all
        ).pack(side="right", padx=5)

    def evaluate_all_answers(self):
        total_score = 0
        max_score = 0
        
        for sheet_name, answer_text in self.answers.items():
            # Get question-specific data
            factors = factors_dict[sheet_name]
            answer_data = ansl_dict[sheet_name]
            
            # Evaluate this answer
            result = self.evaluate_single_answer(
                answer_text,
                answer_data['sample_answers'],
                answer_data['keywords'],
                factors
            )
            
            self.results[sheet_name] = result
            total_score += result['score']
            max_score += 10  # Each question is out of 10
            
            # Update the question frame
            self.update_question_frame(sheet_name, result)
        
        # Update total score
        ttk.Label(
            self.total_frame,
            text=f"Total Score: {total_score:.1f}/{max_score}",
            font=("Helvetica", 14, "bold")
        ).pack()

    def evaluate_single_answer(self, answer_text, sample_answers, keywords, factors):
        """Evaluate a single answer and return metrics."""
        # Initialize metrics
        similarity = 0
        grammar_score = 0
        keyword_score = 0
        keyword_order = 0
        
        # Calculate grammar score
        text_words = answer_text.strip().split()
        english_words = load_words()
        valid_words = 0
        
        if text_words:
            for word in text_words:
                temp = word.lower().rstrip('.')
                if temp in english_words:
                    valid_words += 1
            grammar_score = valid_words / len(text_words)
        
        # Calculate similarity with sample answers
        if valid_words > 7:
            for sample in sample_answers:
                similarity += (fuzz.token_set_ratio(sample, answer_text) + 
                             fuzz.ratio(sample, answer_text))
            similarity = similarity / (len(sample_answers) * 200)  # Normalize to 0-1
        
        # Calculate keyword score
        found_keywords = []
        for word in text_words:
            lw = word.lower()
            if lw in keywords and lw not in found_keywords:
                idx = keywords.index(lw)
                if idx >= 2:
                    keyword_score += 0.05
                elif idx == 0:
                    keyword_score += 0.1
                elif idx == 1:
                    keyword_score += 0.08
                found_keywords.append(lw)
        
        # Calculate keyword order accuracy
        if found_keywords:
            order_check = []
            for i in range(len(found_keywords) - 1):
                if keywords.index(found_keywords[i]) < keywords.index(found_keywords[i + 1]):
                    order_check.extend([found_keywords[i], found_keywords[i + 1]])
            keyword_order = len(order_check) / len(found_keywords) if found_keywords else 0
        
        # Calculate final score
        raw_score = (
            (similarity * factors['frf']) +
            (keyword_score * factors['ktf']) +
            (keyword_order * factors['cmf']) +
            (grammar_score * factors['gmf'])
        )
        
        # Convert to 10-point scale
        score = self.convert_to_ten_point_scale(raw_score)
        
        return {
            'score': score,
            'similarity': similarity,
            'grammar_score': grammar_score,
            'keyword_score': keyword_score,
            'keyword_order': keyword_order,
            'found_keywords': found_keywords
        }

    def convert_to_ten_point_scale(self, r):
        """Convert raw score to 10-point scale."""
        if r > 0.95: return 10
        elif r > 0.9: return 9.5
        elif r > 0.85: return 9
        elif r > 0.8: return 8.5
        elif r > 0.75: return 8
        elif r > 0.7: return 7.5
        elif r > 0.65: return 7
        elif r > 0.6: return 6.5
        elif r > 0.55: return 6
        elif r > 0.5: return 5.5
        elif r > 0.45: return 5
        elif r > 0.4: return 4.5
        elif r > 0.35: return 4
        elif r > 0.3: return 3.5
        elif r > 0.25: return 3
        elif r > 0.2: return 2.5
        elif r > 0.15: return 2
        elif r > 0.1: return 1.5
        elif r > 0.05: return 1
        else: return 0

    def update_question_frame(self, sheet_name, result):
        """Update the display frame for a single question's results."""
        frame = self.question_frames[sheet_name]
        
        ttk.Label(
            frame,
            text=f"Score: {result['score']:.1f}/10",
            font=("Helvetica", 11, "bold")
        ).pack(anchor="w")
        
        ttk.Label(
            frame,
            text=f"Similarity: {result['similarity']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        ttk.Label(
            frame,
            text=f"Grammar Accuracy: {result['grammar_score']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        ttk.Label(
            frame,
            text=f"Keywords Found: {len(result['found_keywords'])}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        ttk.Label(
            frame,
            text=f"Keyword Order Accuracy: {result['keyword_order']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")

    def open_detailed_report(self):
        DetailedReportWindow(self)

    def close_all(self):
        self.master.destroy()


class DetailedReportWindow(tk.Toplevel):
    """Window showing the full detailed evaluation report for all questions."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Detailed Evaluation Report")
        self.geometry("900x700+300+50")
        self.configure(bg="#ffffff")
        
        # Get results from parent
        self.results = parent.results
        self.answers = parent.answers
        
        # Create scrollable frame
        self.create_scrollable_frame()
        self.create_widgets()

    def create_scrollable_frame(self):
        # Create canvas with scrollbar
        self.canvas = tk.Canvas(self)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Pack scrollbar and canvas
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

    def create_widgets(self):
        container = ttk.Frame(self.scrollable_frame, padding="10 10 10 10")
        container.pack(fill="both", expand=True)
        
        # Calculate total score
        total_score = sum(result['score'] for result in self.results.values())
        max_score = len(self.results) * 10
        
        # Display overall score
        ttk.Label(
            container,
            text=f"Total Score: {total_score:.1f}/{max_score}",
            font=("Helvetica", 14, "bold")
        ).pack(anchor="w", pady=10)
        
        # Display detailed results for each question
        for sheet_name in sorted(self.results.keys()):
            self.create_question_section(container, sheet_name)
        
        # Close button
        ttk.Button(
            container,
            text="Close",
            command=self.destroy
        ).pack(pady=10)

    def create_question_section(self, container, sheet_name):
        # Create frame for this question
        question_frame = ttk.LabelFrame(
            container,
            text=f"Question {sheet_name[1]} Details",
            padding="10 10 10 10"
        )
        question_frame.pack(fill="x", pady=10)
        
        # Question text
        ttk.Label(
            question_frame,
            text="Question:",
            font=("Helvetica", 11, "bold")
        ).pack(anchor="w")
        
        ttk.Label(
            question_frame,
            text=Qtext_dict[sheet_name],
            wraplength=800,
            font=("Helvetica", 10)
        ).pack(anchor="w", pady=(0, 10))
        
        # Your answer
        ttk.Label(
            question_frame,
            text="Your Answer:",
            font=("Helvetica", 11, "bold")
        ).pack(anchor="w")
        
        ttk.Label(
            question_frame,
            text=self.answers[sheet_name],
            wraplength=800,
            font=("Helvetica", 10)
        ).pack(anchor="w", pady=(0, 10))
        
        # Detailed metrics
        result = self.results[sheet_name]
        factors = factors_dict[sheet_name]
        
        metrics_frame = ttk.Frame(question_frame)
        metrics_frame.pack(fill="x", pady=5)
        
        # Left column
        left_frame = ttk.Frame(metrics_frame)
        left_frame.pack(side="left", padx=10)
        
        ttk.Label(
            left_frame,
            text=f"Score: {result['score']:.1f}/10",
            font=("Helvetica", 10, "bold")
        ).pack(anchor="w")
        
        ttk.Label(
            left_frame,
            text=f"Similarity Factor ({factors['frf']}): {result['similarity']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        ttk.Label(
            left_frame,
            text=f"Grammar Factor ({factors['gmf']}): {result['grammar_score']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        # Right column
        right_frame = ttk.Frame(metrics_frame)
        right_frame.pack(side="left", padx=10)
        
        ttk.Label(
            right_frame,
            text=f"Keyword Factor ({factors['ktf']}): {result['keyword_score']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        ttk.Label(
            right_frame,
            text=f"Keyword Order Factor ({factors['cmf']}): {result['keyword_order']:.2%}",
            font=("Helvetica", 10)
        ).pack(anchor="w")
        
        # Keywords found
        if result['found_keywords']:
            ttk.Label(
                question_frame,
                text="Keywords Found:",
                font=("Helvetica", 10, "bold")
            ).pack(anchor="w", pady=(10, 0))
            
            ttk.Label(
                question_frame,
                text=", ".join(result['found_keywords']),
                wraplength=800,
                font=("Helvetica", 10)
            ).pack(anchor="w")
        
        # Sample answers
        ttk.Label(
            question_frame,
            text="Sample Answers:",
            font=("Helvetica", 10, "bold")
        ).pack(anchor="w", pady=(10, 0))
        
        for i, answer in enumerate(ansl_dict[sheet_name]['sample_answers'], 1):
            ttk.Label(
                question_frame,
                text=f"{i}. {answer}",
                wraplength=800,
                font=("Helvetica", 10)
            ).pack(anchor="w")


# ---------------- Main Application Launch ----------------
def main():
    root = tk.Tk()

    # Use ttk styles for a modern look
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("TLabel", font=("Helvetica", 11))
    style.configure("TButton", font=("Helvetica", 10, "bold"))
    style.configure("TEntry", font=("Helvetica", 11))
    style.configure("TCombobox", font=("Helvetica", 11))

    app = LoginWindow(root)
    root.mainloop()

if __name__ == '__main__':
    main()

