import os
import fitz  # PyMuPDF
import openpyxl
import google.generativeai as genai
from openpyxl import Workbook
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import shutil
from datetime import datetime
import threading

# STEP 1: Configure Gemini API
genai.configure(api_key='AIzaSyDM5zj3dyNoMILBEYVNy00w8WKRdsLHy9U')
model = genai.GenerativeModel('gemini-1.5-flash')

# Directory to save the Excel file
questions_dir = "/Users/harshporwal/Desktop/MAIN/AI_NLP/Capstone Final/Questions"
os.makedirs(questions_dir, exist_ok=True)

class LoginFrame(ttk.Frame):
    def __init__(self, master, show_main_app_callback):
        super().__init__(master, padding="20")
        self.master = master
        self.show_main_app_callback = show_main_app_callback
        self.grid(sticky=(tk.W, tk.E, tk.N, tk.S))
        self.create_widgets()
        self.center_window(300, 200)

    def create_widgets(self):
        ttk.Label(self, text="Username:").grid(row=0, column=0, pady=5)
        self.username = ttk.Entry(self)
        self.username.grid(row=0, column=1, pady=5)
        
        ttk.Label(self, text="Password:").grid(row=1, column=0, pady=5)
        self.password = ttk.Entry(self, show="*")
        self.password.grid(row=1, column=1, pady=5)
        
        ttk.Button(self, text="Login", command=self.login).grid(row=2, column=0, columnspan=2, pady=20)

    def center_window(self, width, height):
        self.master.update_idletasks()  # update geometry info
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.master.geometry(f"{width}x{height}+{x}+{y}")

    def login(self):
        if self.username.get() == "admin" and self.password.get() == "admin":
            # Hide login frame and show main application
            self.grid_forget()
            self.show_main_app_callback()
        else:
            messagebox.showerror("Error", "Invalid username or password")


class MainAppFrame(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding="20")
        self.master = master
        self.grid(sticky=(tk.W, tk.E, tk.N, tk.S))
        self.create_widgets()
        self.center_window(500, 400)

    def create_widgets(self):
        # Subject Name
        ttk.Label(self, text="Subject Name:").grid(row=0, column=0, pady=10, sticky='w')
        self.subject_name = ttk.Entry(self, width=40)
        self.subject_name.grid(row=0, column=1, pady=10, padx=(10, 0), sticky='ew')
        
        # File Upload: Display the selected PDF path
        ttk.Label(self, text="Selected PDF:").grid(row=1, column=0, pady=10, sticky='w')
        self.file_path = tk.StringVar()
        self.file_label = ttk.Label(self, textvariable=self.file_path, wraplength=300)
        self.file_label.grid(row=1, column=1, pady=10, padx=(10, 0), sticky='ew')

        # Browse Button
        self.browse_button = ttk.Button(self, text="Browse PDF", command=self.browse_file)
        self.browse_button.grid(row=2, column=0, columnspan=2, pady=10)

        # Process Button
        self.process_button = ttk.Button(self, text="Process PDF", command=self.process_file)
        self.process_button.grid(row=3, column=0, columnspan=2, pady=20)

        # Status Label
        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(self, textvariable=self.status_var, wraplength=400)
        self.status_label.grid(row=4, column=0, columnspan=2, pady=10)

        # Configure grid weights
        self.columnconfigure(1, weight=1)

    def center_window(self, width, height):
        self.master.update_idletasks()
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.master.geometry(f"{width}x{height}+{x}+{y}")

    def browse_file(self):
        try:
            filename = filedialog.askopenfilename(
                parent=self.master,
                title="Select PDF File",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            if filename:
                self.file_path.set(filename)
                self.status_var.set("PDF file selected successfully!")
        except Exception as e:
            self.status_var.set(f"Error selecting file: {str(e)}")
            messagebox.showerror("Error", f"Error selecting file: {str(e)}")

    def process_file(self):
        if not self.subject_name.get().strip():
            messagebox.showerror("Error", "Please enter subject name")
            return

        if not self.file_path.get().strip():
            messagebox.showerror("Error", "Please select a PDF file")
            return

        # Disable buttons during processing
        self.process_button.config(state=tk.DISABLED)
        self.browse_button.config(state=tk.DISABLED)
        self.status_var.set("⏳ Processing... please wait.")

        threading.Thread(target=self._process_pdf_thread, daemon=True).start()

    def _process_pdf_thread(self):
        try:
            # Build the Excel file name (no timestamp)
            excel_output = os.path.join(questions_dir, f"{self.subject_name.get()}.xlsx")
            process_pdf_and_create_excel(self.file_path.get(), excel_output)
            self.master.after(0, lambda: self.processing_done(success=True))
        except Exception as e:
            self.master.after(0, lambda: self.processing_done(success=False, error=e))

    def processing_done(self, success, error=None):
        # Re-enable the buttons
        self.process_button.config(state=tk.NORMAL)
        self.browse_button.config(state=tk.NORMAL)

        if success:
            self.status_var.set("✅ Processing completed successfully!")
            messagebox.showinfo("Success", "PDF processed and Excel file created successfully!")
        else:
            err_msg = f"❌ Error during processing: {str(error)}"
            self.status_var.set(err_msg)
            messagebox.showerror("Error", err_msg)

# STEP 2: Read and extract text from PDF
def extract_text_from_pdf(pdf_path):
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"File not found: {pdf_path}")
    
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    doc.close()
    return full_text

# STEP 3: Ask Gemini to extract questions and keywords
def gemini_api_extract(pdf_text):
    prompt = """
You are given the content of a question paper or technical PDF.

Your job is to extract exactly 5 questions: Q1, Q2, Q3, Q4, Q5.

For each question, return the following:

Q1:
question: <full question text>
keywords_explained:
- <Paragraph 1 explaining a key concept or term relevant to the question>
- <Paragraph 2 explaining another concept>
- <Paragraph 3 ...>
- <Paragraph 4 ...>
- <Paragraph 5 ...>

Each bullet point under "keywords_explained" should be a **short paragraph (2-3 lines)** describing a **key concept, keyword, or important term** relevant to the question. These should **explain the meaning and importance** of the term in the context of the question.

DO NOT split questions into subquestions like Q2a/Q2b. Group them under Q2 itself.

Return output in clean JSON format (suitable for parsing using json.loads()).
"""
    response = model.generate_content([prompt, pdf_text])
    try:
        raw_response = response.text.strip()
        if raw_response.startswith("```json"):
            raw_response = raw_response[7:].strip()
        if raw_response.endswith("```"):
            raw_response = raw_response[:-3].strip()
        data = json.loads(raw_response)
        print("Extracted Data:", data)
    except Exception as e:
        print("Error parsing JSON from Gemini response:", e)
        print("Raw Response:\n", response.text)
        raise e
    return data

# STEP 4: Write to Excel
def create_excel_file(data, output_filename="output.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for q_key in ["Q1", "Q2", "Q3", "Q4", "Q5"]:
        sheet = wb.create_sheet(q_key)
        headers = ["Values", "Similarity factor", "Keyword Factor", "Grammar Factor", "Keyword Accuracy Factor"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        for col, val in enumerate([1.0, 0.3, 1,0.2,0.3], 1):
            sheet.cell(row=2, column=col, value=val)
        sheet.cell(row=3, column=1, value="Question:")
        question_text = data.get(q_key, {}).get("question", "Not available")
        sheet.cell(row=4, column=1, value=f"{q_key}: {question_text}")
        sheet.cell(row=6, column=1, value="Keywords:")
        keywords = data.get(q_key, {}).get("keywords_explained", [])
        for i, keyword in enumerate(keywords, start=7):
            sheet.cell(row=i, column=1, value=keyword)
    wb.save(output_filename)
    print(f"✅ Excel file '{output_filename}' created successfully!")

# STEP 5: Main Process
def process_pdf_and_create_excel(pdf_path, excel_output="output.xlsx"):
    pdf_text = extract_text_from_pdf(pdf_path)
    extracted_data = gemini_api_extract(pdf_text)
    create_excel_file(extracted_data, excel_output)

def main():
    root = tk.Tk()
    root.title("Professor Question Uploader (Backend)")
    # Start with the login frame; when login is successful, show the main app frame.
    def show_main_app():
        MainAppFrame(root)
    LoginFrame(root, show_main_app_callback=show_main_app)
    root.mainloop()

if __name__ == "__main__":
    main()
